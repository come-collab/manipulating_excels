import streamlit as st
import openpyxl
from openpyxl import load_workbook
import os
import pandas as pd
from collections import Counter
from datetime import datetime
#Adding the credential for every member of the BDF
import unicodedata


# Define a directory to store the uploaded Excel file globally
EXCEL_FILE_PATH = "shared_excel.xlsx"
PLAYER_LIST_FILE_PATH = "List_of_Player.xlsx"
#Membre de la bdf + Session Invité
USER_CREDENTIALS = {
    #Add to didier la possibilité de dire qui l'a kill aussi en super user
    "Didier" : "",
    "Côme": "",
    "Baptiste":"",
    "Steve":"",
    "Nico":"",
    "Béa":"",
    "Chris":"",
    "Issam":"",
    "Marco":"",
    "Fanny":"",
    "Guigui":"",
    "Gégé":"",
    "Lucas":"",
    "Pat":"",
    "Dylan":"",
    "Maxime":"",
    "Hugo W":"",
    "Marine":"",
    "Yohan":"",
    "Sylvie P":"",
    "Tony":"",
    "Gérard":"",
    "Chloé":"",
    "Pierrot":"",
    "Come":"",
    "Baptiste":"",
    "Flo":"",
    "Sylvie B":"",
    "Gaby":"",
    "Greg":"",
    #Session invité particuliere, tu choisis ton nom
    "Invités":"",

}

st.set_page_config(page_title='Le Site de la BDF',page_icon="../logo_bdf.png")
st.image('../logo_bdf.png')

def report_elimination(elimination_df, player_df, eliminated_player, killer):
    players = player_df["Joueur"].dropna().unique()
    total_players = len(players)

    # Check if eliminated_player has already been eliminated
    if eliminated_player in elimination_df['Joueur'].dropna().values:
        return (elimination_df, False, f"{eliminated_player} a déjà été éliminé.")

    # Count the number of eliminated players
    number_of_eliminated_players = elimination_df['Joueur'].dropna().nunique()

    # Compute the 'Classement' for the eliminated player
    classement = total_players - number_of_eliminated_players

    current_time = datetime.now().strftime("%H:%M")

    # Find the row where 'Classement' equals the calculated classement
    row_index = elimination_df.index[elimination_df['Classement'] == classement].tolist()
    if not row_index:
        return (elimination_df, False, f"Aucune ligne trouvée pour le classement {classement}.")

    # Update the existing row
    index = row_index[0]
    elimination_df.at[index, 'Joueur'] = eliminated_player
    elimination_df.at[index, 'Heure'] = current_time
    elimination_df.at[index, 'Killer'] = killer
    elimination_df.at[index, 'Points'] = None  # Will be updated later

    # Check if only one player remains
    eliminated_players = elimination_df['Joueur'].dropna().unique()
    if len(eliminated_players) == total_players - 1:
        # Assign 'Classement' 1 to the last remaining player
        remaining_players = set(players) - set(eliminated_players)
        if remaining_players:
            last_player = remaining_players.pop()
            # Find the row with 'Classement' == 1
            first_place_index = elimination_df.index[elimination_df['Classement'] == 1].tolist()
            if first_place_index:
                index = first_place_index[0]
                elimination_df.at[index, 'Joueur'] = last_player
                elimination_df.at[index, 'Heure'] = ""  # No elimination time
                elimination_df.at[index, 'Killer'] = ""  # No killer
                elimination_df.at[index, 'Points'] = None  # Will be updated later

    return (elimination_df, True, f"{eliminated_player} a été éliminé (Classement: {classement}, Heure: {current_time})")

def update_points(elimination_df):
    # Define point allocation based on ranking among credentialed users
    points_mapping = {1: 10, 2: 6, 3: 4, 4: 2, 5: 1}
    
    # Ensure 'Classement' is integer
    elimination_df['Classement'] = elimination_df['Classement'].astype(int)
    
    # Normalize 'Joueur' names in the DataFrame
    elimination_df['Normalized_Joueur'] = elimination_df['Joueur'].apply(normalize_name)
    
    # Create a list of normalized credentialed user names
    normalized_credentials = [normalize_name(name) for name in USER_CREDENTIALS.keys()]
    
    # Identify credentialed users
    elimination_df['IsCredentialUser'] = elimination_df['Normalized_Joueur'].isin(normalized_credentials)
    
    # Sort elimination DataFrame by 'Classement'
    elimination_df = elimination_df.sort_values('Classement').reset_index(drop=True)
    
    # Filter credentialed users and get their positions
    credentialed_df = elimination_df[elimination_df['IsCredentialUser']].copy()
    credentialed_df = credentialed_df.reset_index()  # Preserve original indices
    
    # Assign points based on position among credentialed users
    credentialed_df['Credential_Rank'] = range(1, len(credentialed_df) + 1)
    credentialed_df['Points'] = credentialed_df['Credential_Rank'].map(points_mapping).fillna(0).astype(int)
    
    # Update the 'Points' column in the main DataFrame
    elimination_df['Points'] = 0  # Initialize all points to 0
    elimination_df.loc[credentialed_df['index'], 'Points'] = credentialed_df['Points'].values
    
    # Drop temporary columns
    elimination_df = elimination_df.drop(columns=['Normalized_Joueur', 'IsCredentialUser'])
    
    return elimination_df

def normalize_name(name):
    if not isinstance(name, str):
        return ''
    # Remove accents and special characters
    name = unicodedata.normalize('NFD', name).encode('ascii', 'ignore').decode('utf-8')
    # Convert to lowercase and strip whitespace
    return name.strip().lower()


# Function to load the Excel file and return a DataFrame
def load_excel_to_dataframe(file_path):
    if os.path.exists(file_path):
        df = pd.read_excel(file_path, engine='openpyxl')  # Load the Excel file into a DataFrame
        return df
    return None


def load_excel_to_dataframe_2(file_path):
    try:
        df = pd.read_excel(file_path,skiprows=[0])
        return df
    except Exception as e:
        st.error(f"Error loading Excel file: {e}")
        return None
    

def save_dataframe_to_excel(file_path, df, top_row_text="BDF Edition 9 "):
    try:
        # First, save the DataFrame to Excel
        df.to_excel(file_path, index=False)

        # Now, open the workbook and add the top row
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Insert a new row at the top
        ws.insert_rows(1)
        
        # Add the text to cell A1 of the new row
        ws['A1'] = top_row_text

        # Save the modified workbook
        wb.save(file_path)
        
        st.success("Excel file saved successfully with top row inserted.")
    except Exception as e:
        st.error(f"Error saving Excel file: {e}")


def is_credential_user(player_name):
    normalized_player_name = normalize_name(player_name)
    normalized_credentials = [normalize_name(name) for name in USER_CREDENTIALS.keys()]
    return normalized_player_name in normalized_credentials

def login():
    st.title("Login Page")
    username = st.text_input("Utilisateur")

    password = st.text_input("Mot de passe", type="password")
    login_button = st.button("Connexion")

    # Normalize username
    normalized_username = normalize_name(username)

    # Check login credentials
    if login_button:
        normalized_credentials = {normalize_name(k): v for k, v in USER_CREDENTIALS.items()}
        if normalized_username in normalized_credentials and normalized_credentials[normalized_username] == password:
            st.session_state['logged_in'] = True
            st.session_state['username'] = username  # Keep the original username for display
            st.success("Login successful")
            st.rerun()
        else:
            st.error("Invalid username or password")


def modify_excel(workbook, sheet_name, cell, value):
    if workbook and sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet[cell] = value
    return workbook
# Function to save the workbook globally
def save_excel(workbook, file_path):
    workbook.save(file_path)

# Function to load the Excel file using openpyxl
def load_excel_with_openpyxl(file_url=None):
    if os.path.exists(file_url):
        workbook = openpyxl.load_workbook(file_url)
        return workbook
    return None
    

# Function to display a specific page for user1
def user1_page(): 
    st.title("Panneau Admin Didier")
    st.write("Dashboard personnalisé")

    # Upload the first file (Excel to modify)
    uploaded_file = st.file_uploader("Charge l'excel à compléter", type=["xlsx"], key="excel_upload")
    if uploaded_file is not None:
        st.session_state['uploaded_file'] = uploaded_file  # Store in session state temporarily
        if st.button("Submit Excel File"):  # Submit button for the Excel file
            with open(EXCEL_FILE_PATH, "wb") as f:
                f.write(st.session_state['uploaded_file'].getbuffer())  # Save the Excel file
            st.success("Excel file uploaded and saved globally!")
    else:
        st.info("Please upload the Excel file to be modified.")

    # Upload the second file (list of players)
    uploaded_file_list_player = st.file_uploader("Upload la liste des joueurs", type=["xlsx"], key="player_list_upload")
    if uploaded_file_list_player is not None:
        st.session_state['uploaded_file_list_player'] = uploaded_file_list_player
        if st.button("Submit Player List"):
            with open(PLAYER_LIST_FILE_PATH, "wb") as f:
                f.write(st.session_state['uploaded_file_list_player'].getbuffer())
            st.success("List of players file uploaded and saved globally!")
            
            # Load the player list and the shared Excel file
            player_df = load_excel_to_dataframe(PLAYER_LIST_FILE_PATH)
            shared_df = load_excel_to_dataframe_2(EXCEL_FILE_PATH)
            
            if player_df is not None and shared_df is not None:
                # Get the number of players
                num_players = len(player_df)
                
                # Update the Classement column
                shared_df['Classement'] = range(1, num_players + 1)
                
                # Save the updated shared Excel file
                save_dataframe_to_excel(EXCEL_FILE_PATH, shared_df)
                
                st.success(f"Classement column updated with numbers 1 to {num_players}")
            else:
                st.error("Failed to load player list or shared Excel file")
    else:
        st.info("Please upload the list of players file.")

    # Add a new section for Didier to specify his elimination
    st.subheader("Gestion de votre élimination")

    # Check if the PLAYER_LIST file exists
    if not os.path.exists(PLAYER_LIST_FILE_PATH):
        st.error("No PLAYER_LIST Excel file uploaded yet.")
    else:
        player_df = load_excel_to_dataframe(PLAYER_LIST_FILE_PATH)
        
        if player_df is not None and "Joueur" in player_df.columns:
            players = player_df["Joueur"].dropna().unique()
            total_players = len(players)

            # Load the elimination DataFrame
            elimination_df = load_excel_to_dataframe_2(EXCEL_FILE_PATH)
            if elimination_df is None:
                st.error("Failed to load the elimination Excel file.")
                return

            # Determine remaining players
            eliminated_players = elimination_df['Joueur'].dropna().unique()
            remaining_players = set(players) - set(eliminated_players) - {"Didier"}
            remaining_players = list(remaining_players)

            # Display a selectbox for Didier to select the player who eliminated him
            if remaining_players:
                selected_player = st.selectbox("Qui vous a éliminé ?", remaining_players)
            else:
                st.info("Aucun joueur restant pour vous éliminer.")
                return

            if st.button("Confirmer votre élimination"):
                # Reload elimination_df in case it has been updated
                elimination_df = load_excel_to_dataframe_2(EXCEL_FILE_PATH)
                if elimination_df is None:
                    st.error("Failed to load the elimination Excel file.")
                    return

                # Use the report_elimination function
                eliminated_player = "Didier"
                elimination_df, success, message = report_elimination(elimination_df, player_df, eliminated_player, selected_player)
                
                if not success:
                    st.warning(message)
                    return

                # Update points
                updated_df = update_points(elimination_df)

                # Save the updated DataFrame
                save_dataframe_to_excel(EXCEL_FILE_PATH, updated_df)

                st.success(message)
        else:
            st.error("Failed to load the PLAYER_LIST Excel file or 'Joueur' column is missing.")
# Function to display a general page for other users
def general_page():
    st.title(f"Salut, {st.session_state['username']}!")
    st.write("Gestion des éliminations")

    # Load the elimination DataFrame first
    elimination_df = load_excel_to_dataframe_2(EXCEL_FILE_PATH)
    if elimination_df is None:
        st.error("Le fichier d'élimination n'est pas disponible.")
        return

    # Load the player list
    player_df = load_excel_to_dataframe(PLAYER_LIST_FILE_PATH)
    if player_df is None or "Joueur" not in player_df.columns:
        st.error("La liste des joueurs n'est pas disponible.")
        return

    players = player_df["Joueur"].dropna().unique()
    total_players = len(players)

    # Determine remaining players
    eliminated_players = elimination_df['Joueur'].dropna().unique()
    remaining_players = set(players) - set(eliminated_players) - {st.session_state['username']}
    remaining_players = list(remaining_players)

    # Display a selectbox for the user to select the player who eliminated them
    if remaining_players:
        selected_player = st.selectbox("Qui vous a éliminé ?", remaining_players)
    else:
        st.info("Aucun joueur restant pour vous éliminer.")
        return

    current_user = st.session_state['username']

    # Handle user elimination logic
    if st.button("Confirmer l'élimination"):
        # Reload elimination_df in case it has been updated
        elimination_df = load_excel_to_dataframe_2(EXCEL_FILE_PATH)
        if elimination_df is None:
            st.error("Le fichier d'élimination n'est pas disponible.")
            return

        elimination_df, success, message = report_elimination(elimination_df, player_df, current_user, selected_player)

        if not success:
            st.warning(message)
            return

        # Update points
        updated_df = update_points(elimination_df)

        # Save the updated DataFrame
        save_dataframe_to_excel(EXCEL_FILE_PATH, updated_df)

        st.success(message)


def invited_user_page():
    st.title("Page Invités")
    st.write("Sélectionnez votre nom et indiquez qui vous a éliminé")

    # Load the elimination DataFrame first
    elimination_df = load_excel_to_dataframe_2(EXCEL_FILE_PATH)
    if elimination_df is None:
        st.error("Impossible de charger le fichier d'élimination.")
        return

    # Load the player list
    player_df = load_excel_to_dataframe(PLAYER_LIST_FILE_PATH)
    if player_df is None or "Joueur" not in player_df.columns:
        st.error("Impossible de charger la liste des joueurs.")
        return

    all_players = player_df["Joueur"].dropna().unique()
    invited_players = [player for player in all_players if player not in USER_CREDENTIALS]

    if not invited_players:
        st.error("Aucun joueur invité disponible.")
        return

    # Let the invited user select their name
    selected_name = st.selectbox("Sélectionnez votre nom", invited_players)

    # Determine remaining players
    eliminated_players = elimination_df['Joueur'].dropna().unique()
    remaining_players = set(all_players) - set(eliminated_players) - {selected_name}
    remaining_players = list(remaining_players)

    # Let the user select who eliminated them (from remaining players)
    selected_killer = st.selectbox("Qui vous a éliminé ?", remaining_players)

    if st.button("Confirmer l'élimination"):
        # Reload elimination_df in case it has been updated
        elimination_df = load_excel_to_dataframe_2(EXCEL_FILE_PATH)
        if elimination_df is None:
            st.error("Impossible de charger le fichier d'élimination.")
            return

        elimination_df, success, message = report_elimination(elimination_df, player_df, selected_name, selected_killer)

        if not success:
            st.warning(message)
            return

        # Update points
        updated_df = update_points(elimination_df)
        save_dataframe_to_excel(EXCEL_FILE_PATH, updated_df)

        st.success(message)
# Logout function
def logout():
    st.session_state['logged_in'] = False
    st.session_state['username'] = None

# Page selection (authenticated content)
def page_selector():
    # Check if the logged-in user is user1 or another user
    if st.session_state['username'] == "Didier":
        user1_page()  # Show specific content for user1
    elif st.session_state['username'] == "Invités":
        invited_user_page()
    else:
        general_page()  # Show general content for all other users

# Main app
if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False

if st.session_state['logged_in']:
    st.sidebar.title("Navigation")
    page_selector()  # Show user-specific pages based on the logged-in user
    if st.sidebar.button("Logout"):
        logout()
else:
    login()